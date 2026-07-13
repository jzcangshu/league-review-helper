#!/usr/bin/env python3
"""Prepare a new school folder for youth league PDF review."""

from __future__ import annotations

import argparse
import json
import re
import sys
from copy import copy
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string


STATUS_FILE = ".review-status.json"
NAME_HEADERS = {"姓名", "学生姓名", "团员姓名", "人员姓名", "发展对象姓名", "申请人姓名", "成员姓名"}
CONTEXT_HEADERS = ("序号", "编号", "学校", "单位", "班级", "姓名", "性别", "民族", "身份证", "出生", "籍贯", "电话", "宗教", "备注", "问题", "审核")


def clean(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip()


def locale_key(value: str) -> tuple[int, str]:
    return (len(value), value)


def extract_name(stem: str) -> str:
    value = clean(stem)
    value = re.sub(r"^转PDF[_\-\s]*", "", value, flags=re.I)
    value = re.sub(r"pdf", "", value, flags=re.I)
    value = re.sub(r"[_\-\s]*\d{6,}$", "", value)
    value = re.sub(r"^\d{3}\s*班?\s*", "", value)
    for token in ("入团申请书", "入团志愿书", "入团申请", "入团志愿", "申请书", "志愿书", "审核结果"):
        value = value.replace(token, "")
    value = re.sub(r"[\s_\-—－]+", "", value)
    return value.strip()


@dataclass(frozen=True)
class ExcelRoster:
    path: Path
    sheet: str
    header_row: int
    name_col: int
    names: list[str]


def name_header_score(value: str) -> int:
    header = clean(value).replace("（", "").replace("）", "").replace("(", "").replace(")", "")
    if header == "姓名":
        return 120
    if header in NAME_HEADERS:
        return 115
    if header.endswith("姓名") and len(header) <= 12:
        return 105
    if "姓名" in header and len(header) <= 16:
        return 95
    if header in {"名字", "学生", "团员", "成员"}:
        return 60
    return 0


def looks_like_name(value: str) -> bool:
    name = clean(value)
    if not 2 <= len(name) <= 7 or re.search(r"[A-Za-z0-9→?？]", name):
        return False
    if name in {"男", "女", "是", "否", "无", "有", "汉", "汉族", "回族", "满族", "蒙古族", "壮族"}:
        return False
    return bool(re.fullmatch(r"[\u3400-\u9fff·•]+", name))


def combined_headers(sheet, start_row: int, end_row: int) -> list[str]:
    headers: list[str] = []
    for column in range(1, sheet.max_column + 1):
        parts: list[str] = []
        for row in range(start_row, end_row + 1):
            value = clean(sheet.cell(row, column).value)
            if value and value not in parts:
                parts.append(value)
        headers.append("".join(parts))
    return headers


def resolve_name_column(headers: list[str], requested: str | None) -> int | None:
    if requested:
        value = clean(requested)
        if value.isdigit():
            return int(value)
        if re.fullmatch(r"[A-Za-z]+", value):
            return column_index_from_string(value.upper())
        matches = [index + 1 for index, header in enumerate(headers) if clean(header) == value]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise SystemExit(f"姓名列表头不唯一：{requested}。请改用列号或列字母。")
        return None
    scored = [(name_header_score(header), index + 1) for index, header in enumerate(headers)]
    scored.sort(reverse=True)
    return scored[0][1] if scored and scored[0][0] > 0 else None


def find_roster_in_excel(
    path: Path,
    requested_sheet: str | None = None,
    requested_header_row: int | None = None,
    requested_name_column: str | None = None,
) -> ExcelRoster | None:
    workbook = load_workbook(path, read_only=False, data_only=True)
    candidates: list[tuple[float, ExcelRoster]] = []
    for sheet in workbook.worksheets:
        if requested_sheet and sheet.title != requested_sheet:
            continue
        rows = [requested_header_row] if requested_header_row else range(1, min(sheet.max_row, 80) + 1)
        for row in rows:
            if row is None or row < 1 or row > sheet.max_row:
                continue
            for span in range(1, min(3, row) + 1):
                headers = combined_headers(sheet, row - span + 1, row)
                name_col = resolve_name_column(headers, requested_name_column)
                if not name_col:
                    continue
                names = [
                    clean(sheet.cell(current_row, name_col).value)
                    for current_row in range(row + 1, sheet.max_row + 1)
                    if clean(sheet.cell(current_row, name_col).value)
                ]
                if not names:
                    continue
                likely = sum(looks_like_name(name) for name in names)
                header_score = name_header_score(headers[name_col - 1])
                context = sum(any(token in header for header in headers) for token in CONTEXT_HEADERS)
                score = header_score * 2 + (likely / len(names)) * 55 + context * 12 + min(len(names), 20)
                candidates.append((score, ExcelRoster(path, sheet.title, row, name_col, names)))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    roster = candidates[0][1]
    suspicious = [name for name in roster.names if not looks_like_name(name)]
    if suspicious:
        names = "、".join(suspicious[:8])
        raise SystemExit(f"Excel 姓名列含可疑内容：{names}。请在 WebUI 预览中修正，或修正源 Excel 后重试。")
    return roster


def discover_excel(
    workspace: Path,
    school_dir: Path | None,
    explicit: Path | None,
    requested_sheet: str | None = None,
    requested_header_row: int | None = None,
    requested_name_column: str | None = None,
) -> ExcelRoster:
    if explicit:
        roster = find_roster_in_excel(explicit, requested_sheet, requested_header_row, requested_name_column)
        if not roster:
            raise SystemExit(f"未在 Excel 中找到姓名列：{explicit}")
        return roster

    search_root = school_dir or workspace
    candidates: list[ExcelRoster] = []
    for path in search_root.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        roster = find_roster_in_excel(path, requested_sheet, requested_header_row, requested_name_column)
        if roster and roster.names:
            candidates.append(roster)
    if not candidates:
        raise SystemExit(f"未找到带姓名列的 Excel：{search_root}")
    candidates.sort(key=lambda item: len(item.names), reverse=True)
    return candidates[0]


def discover_pdf_dir(school_dir: Path | None, explicit: Path | None) -> Path:
    if explicit:
        return explicit
    if not school_dir:
        raise SystemExit("未提供 --pdf-dir 时必须提供 --school-dir。")

    candidates: list[tuple[int, Path]] = []
    for directory in [school_dir, *[path for path in school_dir.rglob("*") if path.is_dir()]]:
        count = len(list(directory.glob("*.pdf")))
        if count:
            candidates.append((count, directory))
    if not candidates:
        raise SystemExit(f"未找到 PDF 资料目录：{school_dir}")
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def suspicious_pairs(only_excel: list[str], only_pdf: list[str]) -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    used_pdf: set[str] = set()
    for excel_name in only_excel:
        candidates: list[tuple[float, int, str]] = []
        for pdf_name in only_pdf:
            if pdf_name in used_pdf or len(excel_name) != len(pdf_name):
                continue
            same_positions = sum(1 for left, right in zip(excel_name, pdf_name) if left == right)
            if excel_name[0] == pdf_name[0] and same_positions == len(excel_name) - 1:
                ratio = SequenceMatcher(None, excel_name, pdf_name).ratio()
                candidates.append((ratio, same_positions, pdf_name))
        if candidates:
            candidates.sort(reverse=True)
            pdf_name = candidates[0][2]
            used_pdf.add(pdf_name)
            pairs.append((excel_name, pdf_name))
    return pairs


def is_confident_name_match(excel_name: str, result_name_value: str) -> bool:
    if excel_name == result_name_value:
        return True
    if not excel_name or not result_name_value or len(excel_name) != len(result_name_value):
        return False
    if excel_name[0] != result_name_value[0]:
        return False
    differing_positions = sum(1 for left, right in zip(excel_name, result_name_value) if left != right)
    if differing_positions != 1:
        return False
    return True


def find_confident_result_match(excel_name: str, results: dict[str, str], aliases: dict[str, str]) -> tuple[str | None, str]:
    if excel_name in aliases:
        return aliases[excel_name], "alias"
    if excel_name in results:
        return excel_name, "exact"

    candidates = [name for name in results if is_confident_name_match(excel_name, name)]
    if len(candidates) == 1:
        return candidates[0], "fuzzy"
    if len(candidates) > 1:
        return None, "ambiguous:" + "、".join(sorted(candidates, key=locale_key))
    return None, "missing"


def ensure_empty_file(path: Path) -> bool:
    if path.exists():
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("", encoding="utf-8")
    return True


def update_web_sources(workspace: Path, school: str, pdf_dir: Path) -> None:
    sources_path = workspace / "review-web" / "sources.json"
    if sources_path.exists():
        sources = json.loads(sources_path.read_text(encoding="utf-8"))
        if not isinstance(sources, list):
            sources = []
    else:
        sources = []

    relative_pdf_dir = str(pdf_dir.relative_to(workspace))
    next_sources = [
        source
        for source in sources
        if not (
            isinstance(source, dict)
            and (
                source.get("school") == school
                or str(source.get("folderRelativePath", "")).lower() == relative_pdf_dir.lower()
            )
        )
    ]
    next_sources.append({"school": school, "folderRelativePath": relative_pdf_dir})
    sources_path.parent.mkdir(parents=True, exist_ok=True)
    sources_path.write_text(json.dumps(next_sources, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def parse_aliases(values: list[str]) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise SystemExit(f"别名格式应为 Excel姓名=审核结果姓名：{value}")
        excel_name, review_name = value.split("=", 1)
        aliases[clean(excel_name)] = clean(review_name)
    return aliases


def result_name(path: Path) -> str:
    stem = path.stem
    if stem.endswith("_审核结果"):
        stem = stem[: -len("_审核结果")]
    return extract_name(stem)


def load_review_results(review_dir: Path) -> dict[str, str]:
    if not review_dir.exists():
        return {}
    results: dict[str, str] = {}
    for path in sorted(review_dir.glob("*_审核结果.txt"), key=lambda item: item.name):
        results[result_name(path)] = path.read_text(encoding="utf-8-sig").strip()
    return results


def load_reviewed_names(review_dir: Path) -> set[str]:
    status_path = review_dir / STATUS_FILE
    if not status_path.exists():
        return set()
    try:
        payload = json.loads(status_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return set()
    reviewed = payload.get("reviewed", {})
    return {clean(name) for name in reviewed if clean(name)} if isinstance(reviewed, dict) else set()


def find_result_column(sheet, header_row: int, requested_column: str | None = None) -> tuple[int, str, str]:
    headers = [
        (column, clean(sheet.cell(header_row, column).value))
        for column in range(1, sheet.max_column + 1)
        if clean(sheet.cell(header_row, column).value)
    ]

    if requested_column:
        requested = clean(requested_column)
        for column, header in headers:
            if header == requested:
                return column, header, "specified"
        raise SystemExit(f"未找到指定回写列：{requested_column}。请确认表头行存在该列；脚本不会自动创建回写列。")

    match_rules = [
        ("问题备注", lambda header: header == "问题备注"),
        ("问题", lambda header: header == "问题"),
        ("包含问题", lambda header: "问题" in header),
        ("审核意见", lambda header: header in {"审核意见", "审核结果", "审核备注"}),
        ("备注", lambda header: header == "备注"),
    ]
    for label, predicate in match_rules:
        matches = [(column, header) for column, header in headers if predicate(header)]
        if len(matches) == 1:
            column, header = matches[0]
            return column, header, f"auto:{label}"
        if len(matches) > 1:
            names = "、".join(header for _, header in matches)
            raise SystemExit(f"回写列匹配到多个“{label}”候选：{names}。请用 --result-column 指定准确列名。")

    raise SystemExit("未找到回写列。脚本会依次匹配问题备注、问题、唯一含问题列、审核意见别名和精确备注，不会自动创建新列。")


def red_font_from(cell) -> Font:
    font = copy(cell.font)
    font.color = "FF0000"
    return font


def normal_font_from(cell) -> Font:
    font = copy(cell.font)
    font.color = None
    return font


def pending_font_from(cell) -> Font:
    font = copy(cell.font)
    font.color = "C65911"
    return font


def write_reviews_to_excel(
    roster: ExcelRoster,
    review_dir: Path,
    aliases: dict[str, str],
    missing_text: str,
    pending_text: str,
    result_column: str | None,
) -> dict[str, object]:
    results = load_review_results(review_dir)
    reviewed_names = load_reviewed_names(review_dir)
    workbook = load_workbook(roster.path)
    sheet = workbook[roster.sheet]
    result_col, result_header, result_match = find_result_column(sheet, roster.header_row, result_column)

    written = 0
    blank = 0
    pending: list[str] = []
    missing: list[str] = []
    ambiguous: list[tuple[str, str]] = []
    alias_used: list[tuple[str, str]] = []
    fuzzy_used: list[tuple[str, str]] = []
    used_results: set[str] = set()

    for row in range(roster.header_row + 1, sheet.max_row + 1):
        excel_name = clean(sheet.cell(row, roster.name_col).value)
        if not excel_name:
            continue
        source_name, match_kind = find_confident_result_match(excel_name, results, aliases)
        cell = sheet.cell(row, result_col)

        if source_name and source_name in results:
            content = results[source_name]
            used_results.add(source_name)
            if content:
                cell.value = content
                cell.font = normal_font_from(cell)
                written += 1
            elif source_name in reviewed_names:
                cell.value = None
                cell.font = normal_font_from(cell)
                blank += 1
            else:
                cell.value = pending_text
                cell.font = pending_font_from(cell)
                pending.append(excel_name)
            if match_kind == "alias":
                alias_used.append((excel_name, source_name))
            elif match_kind == "fuzzy":
                fuzzy_used.append((excel_name, source_name))
        else:
            cell.value = missing_text
            cell.font = red_font_from(cell)
            if match_kind.startswith("ambiguous:"):
                ambiguous.append((excel_name, match_kind.split(":", 1)[1]))
            else:
                missing.append(excel_name)

    workbook.save(roster.path)
    unused_results = sorted(set(results) - used_results, key=locale_key)
    return {
        "result_col": result_col,
        "result_header": result_header,
        "result_match": result_match,
        "written": written,
        "blank": blank,
        "pending": pending,
        "missing": missing,
        "ambiguous": ambiguous,
        "alias_used": alias_used,
        "fuzzy_used": fuzzy_used,
        "unused_results": unused_results,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workspace", default=".", help="workspace root, default: current directory")
    parser.add_argument("--school", help="school display name")
    parser.add_argument("--school-dir", help="school root folder; used for auto-discovery")
    parser.add_argument("--excel", help="explicit roster Excel path")
    parser.add_argument("--sheet", help="worksheet name when automatic selection is uncertain")
    parser.add_argument("--header-row", type=int, help="last header row number")
    parser.add_argument("--name-column", help="name column number, letter, or exact header")
    parser.add_argument("--pdf-dir", help="explicit PDF folder")
    parser.add_argument("--review-root", default="审核结果", help="review result root folder")
    parser.add_argument("--update-web-sources", action="store_true", help="register folder in review-web/sources.json")
    parser.add_argument("--write-excel", action="store_true", help="write completed review TXT results into the roster Excel")
    parser.add_argument("--alias", action="append", default=[], help="confirmed name mapping, format: Excel姓名=审核结果姓名")
    parser.add_argument("--missing-text", default="无资料", help="text written in red when no matching review result exists")
    parser.add_argument("--pending-text", default="未审核", help="text written when an empty TXT has not been explicitly reviewed")
    parser.add_argument("--result-column", help="existing Excel header to receive review results; must match the cleaned header exactly")
    args = parser.parse_args()

    workspace = Path(args.workspace).resolve()
    school_dir = (workspace / args.school_dir).resolve() if args.school_dir else None
    excel_path = (workspace / args.excel).resolve() if args.excel else None
    pdf_dir_arg = (workspace / args.pdf_dir).resolve() if args.pdf_dir else None

    roster = discover_excel(
        workspace, school_dir, excel_path, args.sheet, args.header_row, args.name_column
    )
    school = args.school or (school_dir.name if school_dir else roster.path.parent.name)
    review_dir = (workspace / args.review_root / school).resolve()

    if args.write_excel:
        aliases = parse_aliases(args.alias)
        write_report = write_reviews_to_excel(
            roster, review_dir, aliases, args.missing_text, args.pending_text, args.result_column
        )
        print(f"# {school} 审核结果回写报告")
        print()
        print(f"- Excel: {roster.path.relative_to(workspace)}")
        print(f"- 工作表: {roster.sheet}，姓名表头行: {roster.header_row}")
        print(f"- 回写列: {write_report['result_header']}（第 {write_report['result_col']} 列，{write_report['result_match']}）")
        print(f"- 审核结果目录: {review_dir.relative_to(workspace)}")
        print(f"- 写入非空审核结果: {write_report['written']}；写入空结果: {write_report['blank']}")
        print(f"- 未审核空 TXT 标记: {len(write_report['pending'])}")
        print(f"- 无资料红字标记: {len(write_report['missing'])}")
        alias_used = write_report["alias_used"]
        fuzzy_used = write_report["fuzzy_used"]
        mismatch_used = [*alias_used, *fuzzy_used]
        print("- 姓名字形不一致写入: " + ("、".join(f"{left}←{right}" for left, right in mismatch_used) if mismatch_used else "无"))
        if alias_used:
            print("- 使用确认别名: " + "、".join(f"{left}←{right}" for left, right in alias_used))
        if fuzzy_used:
            print("- 高置信模糊写入: " + "、".join(f"{left}←{right}" for left, right in fuzzy_used))
        ambiguous = write_report["ambiguous"]
        if ambiguous:
            print("- 模糊匹配多候选，已按无资料标记: " + "、".join(f"{left}←{right}" for left, right in ambiguous))
        missing = write_report["missing"]
        pending = write_report["pending"]
        print("- 未审核名单: " + ("、".join(pending) if pending else "无"))
        print("- 无资料名单: " + ("、".join(missing) if missing else "无"))
        unused_results = write_report["unused_results"]
        print("- 未写入的审核结果TXT: " + ("、".join(unused_results) if unused_results else "无"))
        return 0

    pdf_dir = discover_pdf_dir(school_dir, pdf_dir_arg)
    pdf_files = sorted(pdf_dir.glob("*.pdf"), key=lambda path: path.name)
    pdf_names = [extract_name(path.stem) for path in pdf_files]
    pdf_name_set = set(pdf_names)
    excel_name_set = set(roster.names)

    created_reviews = 0
    existing_reviews = 0
    for pdf_path, student_name in zip(pdf_files, pdf_names):
        if not student_name:
            continue
        review_path = review_dir / f"{student_name}_审核结果.txt"
        if ensure_empty_file(review_path):
            created_reviews += 1
        else:
            existing_reviews += 1

    if args.update_web_sources:
        update_web_sources(workspace, school, pdf_dir)

    only_excel = [name for name in roster.names if name not in pdf_name_set]
    only_pdf = sorted(pdf_name_set - excel_name_set, key=locale_key)
    pairs = suspicious_pairs(only_excel, only_pdf)

    print(f"# {school} 审核前准备报告")
    print()
    print(f"- Excel: {roster.path.relative_to(workspace)}")
    print(f"- 工作表: {roster.sheet}，姓名表头行: {roster.header_row}")
    print(f"- PDF目录: {pdf_dir.relative_to(workspace)}")
    print(f"- 审核结果目录: {review_dir.relative_to(workspace)}")
    print(f"- 名单人数: {len(roster.names)}；PDF人数: {len(pdf_names)}")
    print(f"- 新建审核结果TXT: {created_reviews}；已有: {existing_reviews}")
    if args.update_web_sources:
        print("- review-web导入配置: 已更新")
    print()
    print("## 核对结果")
    print()
    print("只在名单出现：" + ("、".join(only_excel) if only_excel else "无"))
    print()
    print("只在资料出现：" + ("、".join(only_pdf) if only_pdf else "无"))
    print()
    print("疑似姓名字形不一致：" + ("、".join(f"{left}/{right}" for left, right in pairs) if pairs else "无"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
